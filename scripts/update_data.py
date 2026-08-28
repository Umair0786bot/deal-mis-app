"""Incremental data refresh for the Deal MIS (no raw JSON folder needed).

  python scripts/update_data.py [--downloads DIR] [--from YYYY-MM-DD] [--today YYYY-MM-DD]

What it does
  1. Sellerboard  - every `Ecotero_Dashboard_Products_Group_by_ASIN_<d>_<m>_<y>-<same>_(<stamp>)_sellerboard.com.xlsx`
                    in the downloads folder whose day is >= --from. Latest pull per day wins (Sellerboard restates),
                    and a day with fewer than 200 ASIN rows is rejected as a broken export.
                    Column map (verified against the stored payload to the cent):
                    [dateIdx, asinIdx, Units, Sales, -Sponsored products, -Ads, Net profit, BSR, -Amazon fees, -Cost of Goods, -Refund cost]
                    money in cents; 'Refund сost' is spelt with a Cyrillic es in the export.
  2. Allocations  - the newest `Deal Tracker Updated*.xlsx` Deal Allocation tab replaces data/allocations.js.
                    Tracker deal ids that collide with the reconciled calendar are remapped by (tag, type, start).
  3. Planner rows - `deal-allocation-<TAG>-<date>.csv` SKU DETAIL sections become planner rows for the deal
                    of that tag/type whose window the CSV was pulled for (matched by tag + next upcoming deal).
  4. settings     - built / today.
"""
import argparse, glob, json, os, re, sys, csv, datetime
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'data')
ap = argparse.ArgumentParser()
ap.add_argument('--downloads', default=os.path.expanduser('~/Downloads'))
ap.add_argument('--from', dest='from_', default=None)
ap.add_argument('--today', default=datetime.date.today().isoformat())
A = ap.parse_args()


def load(name):
    s = open(os.path.join(DATA, name + '.js'), encoding='utf-8').read()
    return json.loads(s[s.index('=', 20) + 1:].rstrip().rstrip(';'))


def emit(name, obj):
    s = json.dumps(obj, separators=(',', ':'), ensure_ascii=False)
    open(os.path.join(DATA, name + '.js'), 'w', encoding='utf-8').write('window.DCC=window.DCC||{};window.DCC.%s=%s;\n' % (name, s))
    print('%-16s %7.0f KB' % (name + '.js', len(s) / 1024))


SB = load('sellerboard'); SK = load('skus'); DE = load('deals'); ST = load('settings'); AL = load('allocations'); PL = load('planner')
sku_tag = {s['sku']: s['tag'] for s in SK}
asin_idx = {a[0]: i for i, a in enumerate(SB['asins'])}
from_ = A.from_ or SB['dates'][-1]
print('feed ends', SB['dates'][-1], '- refreshing from', from_)

# ---------- 1. Sellerboard ----------
pat = re.compile(r'Group_by_ASIN_(\d\d)_(\d\d)_(\d{4})-(\d\d)_(\d\d)_(\d{4})_\((\d{4}_\d\d_\d\d_\d\d_\d\d_\d\d_\d+)\)_sellerboard\.com\.xlsx$')
files = {}
for f in glob.glob(os.path.join(A.downloads, 'Ecotero_Dashboard_Products_Group_by_ASIN_*.xlsx')):
    m = pat.search(os.path.basename(f))
    if not m or (m.group(1), m.group(2), m.group(3)) != (m.group(4), m.group(5), m.group(6)):
        continue  # only single-day exports
    day = '%s-%s-%s' % (m.group(3), m.group(2), m.group(1))
    if day < from_:
        continue
    if day not in files or m.group(7) > files[day][0]:
        files[day] = (m.group(7), f)
cents = lambda v: int(round(float(v or 0) * 100))
for day in sorted(files):
    stamp, f = files[day]
    ws = openpyxl.load_workbook(f, read_only=True, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True)); hdr = list(rows[0]); hi = {h: i for i, h in enumerate(hdr)}
    refund_col = next(h for h in hdr if h and h.startswith('Refund') and 'ost' in h)
    body = [r for r in rows[1:] if r[hi['ASIN']]]
    if len(body) < 200:
        print('REJECT', day, 'only', len(body), 'ASIN rows (broken export)'); continue
    if day not in SB['dates']:
        SB['dates'].append(day); SB['dates'].sort()
        order = {d: i for i, d in enumerate(SB['dates'])}
        # dates list changed: re-index every stored row's date index
        old_dates = [d for d in SB['dates'] if d != day]
        remap = {i: order[d] for i, d in enumerate(old_dates)}
        for r in SB['rows']: r[0] = remap[r[0]]
    di = SB['dates'].index(day)
    SB['rows'] = [r for r in SB['rows'] if r[0] != di]
    added = 0; new_asins = 0
    for r in body:
        asin = r[hi['ASIN']]; sku = r[hi['SKU']] or asin
        if asin not in asin_idx:
            SB['asins'].append([asin, sku, sku_tag.get(sku, '?')]); asin_idx[asin] = len(SB['asins']) - 1; new_asins += 1
        g = lambda name: r[hi[name]] if name in hi else 0
        bsr = g('BSR'); bsr = int(bsr) if isinstance(bsr, (int, float)) else 0
        SB['rows'].append([di, asin_idx[asin], int(g('Units') or 0), cents(g('Sales')), -cents(g('Sponsored products (PPC)')), -cents(g('Ads')), cents(g('Net profit')), bsr, -cents(g('Amazon fees')), -cents(g('Cost of Goods')), -cents(g(refund_col))])
        added += 1
    bad = sum(1 for r in SB['rows'] if r[0] == di and abs((r[3] - r[5] - r[8] - r[9] - r[10]) - r[6]) > 1)
    print('Sellerboard %s <- %s rows (%d new ASINs, identity mismatches %d) from pull %s' % (day, added, new_asins, bad, stamp))
SB['rows'].sort(key=lambda r: (r[0], r[1]))
emit('sellerboard', SB)

# ---------- 2. Allocations from the newest tracker ----------
trackers = sorted(glob.glob(os.path.join(A.downloads, 'Deal Tracker Updated*.xlsx')), key=os.path.getmtime)
by_key = {}
for d in DE['rows']:
    by_key.setdefault((d['tag'], d['type'], d['start']), d['id'])
if trackers:
    tr = trackers[-1]; print('tracker', os.path.basename(tr))
    wb = openpyxl.load_workbook(tr, read_only=True, data_only=True)
    cal = {}
    for r in wb['Deal Calendar'].iter_rows(values_only=True):
        if r and isinstance(r[0], str) and re.match(r'^D\d{3}$', r[0]) and hasattr(r[3], 'date'):
            cal[r[0]] = (r[1], r[2], r[3].date().isoformat())
    idmap = {}
    for tid, key in cal.items():
        mine = by_key.get(key)
        if mine and mine != tid:
            idmap[tid] = mine
        elif not mine:
            print('  tracker deal not in the reconciled calendar:', tid, key)
    if idmap: print('  id remap (tracker -> app):', idmap)
    rows = []
    for r in wb['Deal Allocation'].iter_rows(values_only=True):
        if r and isinstance(r[0], str) and re.match(r'^D\d{3}$', r[0]) and r[1]:
            rows.append({'deal': idmap.get(r[0], r[0]), 'sku': str(r[1]).strip(), 'alloc': float(r[2] or 0), 'notes': r[3]})
    if rows:
        AL = rows; emit('allocations', AL)
        from collections import Counter
        print('  allocation rows by deal:', dict(Counter(a['deal'] for a in AL)))

# ---------- 3. Planner rows from deal-allocation CSVs ----------
pcsv = sorted(glob.glob(os.path.join(A.downloads, 'deal-allocation-*-*.csv')))
latest = {}
for f in pcsv:
    m = re.search(r'deal-allocation-([A-Z0-9]+)-(\d{4}-\d\d-\d\d)\.csv$', os.path.basename(f))
    if m and (m.group(1) not in latest or m.group(2) > latest[m.group(1)][0]):
        latest[m.group(1)] = (m.group(2), f)
alloc_by_deal = {}
for a in AL: alloc_by_deal.setdefault(a['deal'], set()).add(a['sku'])
for tag, (pulled, f) in latest.items():
    text = open(f, encoding='utf-8-sig').read()
    if 'SKU DETAIL' not in text: continue
    detail = text.split('SKU DETAIL', 1)[1].strip().splitlines()
    rd = list(csv.DictReader(detail))
    skus = {r['SKU'] for r in rd}
    # the deal this pull belongs to: same tag, not ended at pull date, best SKU overlap with its allocation rows
    cands = [d for d in DE['rows'] if d['tag'] == tag and d['end'] >= pulled]
    if not cands: print('  no deal for planner csv', os.path.basename(f)); continue
    best = max(cands, key=lambda d: (len(skus & alloc_by_deal.get(d['id'], set())), -abs((datetime.date.fromisoformat(d['start']) - datetime.date.fromisoformat(pulled)).days)))
    num = lambda v: None if v in (None, '', '-') else float(str(v).replace(',', ''))
    def col(r, *names):
        for n in names:
            if n in r and r[n] not in (None, ''): return r[n]
        return None
    new = []
    for r in rd:
        new.append({'sku': r['SKU'], 'product': col(r, 'Product'), 'mkt': col(r, 'Marketplace') or 'US', 'fba_now': num(col(r, 'FBA Now', 'FBA On Hand', 'On Hand')), 'vel': num(col(r, 'Velocity/day', 'Velocity', 'Vel/day')), 'min_doh': num(col(r, 'Min DOH')), 'hard_floor': num(col(r, 'Hard Floor')), 'soft_target': num(col(r, 'Soft Target')), 'service': col(r, 'Service Level') or 'p80', 'baseline_demand': num(col(r, 'Expected Units (deal window)', 'Baseline Demand', 'Expected Units')), 'pipeline': num(col(r, 'Pipeline In', 'Pipeline')), 'safe_alloc': num(col(r, 'Safe Allocation', 'Safe')), 'upside': num(col(r, 'Upside', 'Upside (unconfirmed PO)')), 'rec_date': col(r, 'Recovery Date'), 'rec_po': col(r, 'Recovery PO'), 'rec_doh': num(col(r, 'Recovered DOH')), 'post_doh_max': num(col(r, 'Post-Deal DOH (max safe)', 'Post-Deal DOH')), 'post_doh_bal': num(col(r, 'Post-Deal DOH (balanced)', 'Post-Deal DOH')), 'status': col(r, 'Status') or 'SAFE', 'deal': best['id']})
    PL = [p for p in PL if p['deal'] != best['id']] + new
    print('  planner %s (%s) -> %s %s %s: %d rows' % (os.path.basename(f), pulled, best['id'], best['tag'], best['start'], len(new)))
emit('planner', PL)

# ---------- 4. settings ----------
ST['built'] = A.today; ST['today'] = A.today
emit('settings', ST)
print('done: feed now', SB['dates'][0], '-', SB['dates'][-1], len(SB['dates']), 'days,', len(SB['asins']), 'ASINs,', len(SB['rows']), 'rows; today =', A.today)
