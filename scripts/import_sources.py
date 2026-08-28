"""Import the reference tabs the Deals Financials model reads, from the newest workbooks in Downloads:

  Deal Tracker Updated*.xlsx      -> Deal Dashboard (clean columns) => data/dashboard.js
                                     Cost Master (price / FBA / COGS)  => data/skus.js (add or update)
  Deal Analysis Final*.xlsx       -> Cost Master + Updated COGS (l30 units, FBA on hand) => data/skus.js
                                     Historical Deal price            => data/price_history.js
                                     Estimated Upcoming Month charge  => data/sellerboard.js  age{} (AIS units + charge)
  Sellerboard ASINs with no tag get their tag from the updated cost master.

  python scripts/import_sources.py [--downloads DIR] [--since 2026-06-01]
"""
import argparse, glob, json, os, re, datetime
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__))); DATA = os.path.join(ROOT, 'data')
ap = argparse.ArgumentParser(); ap.add_argument('--downloads', default=os.path.expanduser('~/Downloads')); ap.add_argument('--since', default='2026-06-01'); A = ap.parse_args()


def load(name):
    s = open(os.path.join(DATA, name + '.js'), encoding='utf-8').read(); return json.loads(s[s.index('=', 20) + 1:].rstrip().rstrip(';'))


def emit(name, obj):
    s = json.dumps(obj, separators=(',', ':'), ensure_ascii=False)
    open(os.path.join(DATA, name + '.js'), 'w', encoding='utf-8').write('window.DCC=window.DCC||{};window.DCC.%s=%s;\n' % (name, s)); print('%-16s %7.0f KB' % (name + '.js', len(s) / 1024))


def newest(pattern):
    f = sorted(glob.glob(os.path.join(A.downloads, pattern)), key=os.path.getmtime); return f[-1] if f else None


def rows(ws, minrow=1): return [r for r in ws.iter_rows(values_only=True, min_row=minrow) if r and any(x is not None for x in r)]
def num(v):
    if v is None or v == '' or v == '-': return None
    try: return float(str(v).replace('$', '').replace(',', '').replace('%', ''))
    except ValueError: return None
def s(v): return None if v is None else str(v).strip()
ASIN_RE = re.compile(r'^B0[A-Z0-9]{8}$')
def asin_sku(a, b):
    """Cost Master rows sometimes carry SKU and ASIN swapped - decide by the ASIN pattern."""
    a, b = s(a), s(b)
    if b and ASIN_RE.match(b) and not (a and ASIN_RE.match(a)): return b, a
    return a, b
def dstr(v): return v.date().isoformat() if hasattr(v, 'date') else (str(v)[:10] if v else None)


SK = load('skus'); SB = load('sellerboard'); PH = load('price_history')
by_sku = {x['sku']: x for x in SK}; by_asin = {x['asin']: x for x in SK if x.get('asin')}
added = updated = 0

def upsert_cost(tag, asin, sku, price, fba, cogs, extra=None, fill_only=False):
    """fill_only: only set price/fba/cogs where the row has none (Updated COGS is a fallback, Cost Master wins)."""
    global added, updated
    if not sku: return
    row = by_sku.get(sku) or (by_asin.get(asin) if asin else None)
    if not row:
        row = {'sku': sku, 'asin': asin, 'tag': tag, 'price': price, 'fba': fba, 'cogs': cogs, 'brand': None, 'product': None, 'size': None, 'color': None, 'cogs_basis': 'cost master'}
        SK.append(row); by_sku[sku] = row
        if asin: by_asin[asin] = row
        added += 1
    else:
        ch = False
        for k, v in (('price', price), ('fba', fba), ('cogs', cogs)):
            if v is not None and row.get(k) != v and not (fill_only and row.get(k) is not None): row[k] = v; ch = True
        if tag and not row.get('tag'): row['tag'] = tag; ch = True
        if asin and not row.get('asin'): row['asin'] = asin; by_asin[asin] = row; ch = True
        updated += ch
    if extra:
        for k, v in extra.items():
            if v is not None: row[k] = v


# ---------- Deal Analysis file ----------
da = newest('Deal Analysis Final*.xlsx')
if da:
    print('Deal Analysis:', os.path.basename(da)); wa = openpyxl.load_workbook(da, read_only=True, data_only=True)
    for r in rows(wa['Cost Master'], 5):
        if r[1] and r[2] and num(r[3]) is not None: asin, sku = asin_sku(r[1], r[2]); upsert_cost(s(r[0]), asin, sku, num(r[3]), num(r[4]), num(r[5]))
    hdr = None
    for r in rows(wa['Updated COGS']):
        if hdr is None: hdr = [s(x) for x in r]; continue
        d = dict(zip(hdr, r)); sku = s(d.get('sku'))
        if not sku: continue
        upsert_cost(None, None, sku, None, num(d.get('expected-fulfillment-fee-per-unit')), num(d.get('cogs_per_unit')), fill_only=True, extra={'brand': s(d.get('brand')), 'product': s(d.get('product')), 'size': s(d.get('size')), 'color': s(d.get('color')), 'l30': num(d.get('l30_units_sold')), 'fba_on_hand': num(d.get('fba_on_hand_units_LIVE')), 'stock_asof': dstr(d.get('as_of_date')), 'cogs_basis': s(d.get('cogs_basis'))})
    ph_by = {p.get('sku'): p for p in PH if p.get('sku')}; nph = 0
    for r in rows(wa['Historical Deal price']):
        sku, price = s(r[0]), num(r[1]) if len(r) > 1 else None
        if not sku or price is None or sku == 'SKU': continue
        row = by_sku.get(sku); p = ph_by.get(sku)
        if not p: p = {'tag': row['tag'] if row else None, 'size': row.get('size') if row else None, 'asin': row['asin'] if row else None, 'sku': sku, 'price': row.get('price') if row else None, 'deal_price': price, 'reset': None}; PH.append(p); ph_by[sku] = p; nph += 1
        elif p.get('deal_price') is None: p['deal_price'] = price; nph += 1
    print('  historical deal prices added/filled:', nph)
    # AIS: age[asinIdx] = [charge/mo, aged units, in30, in60, in90, rate]
    asin_idx = {a[0]: i for i, a in enumerate(SB['asins'])}; nais = 0
    for r in rows(wa['Estimated Upcoming Month charge']):
        if len(r) < 6 or not r[2] or r[2] == 'Asin': continue
        asin, sku, charge, units = s(r[2]), s(r[3]), num(r[4]), num(r[5])
        if charge is None or units is None: continue
        if asin not in asin_idx:
            row = by_sku.get(sku); SB['asins'].append([asin, sku or asin, (row or {}).get('tag') or s(r[0]) or '?']); asin_idx[asin] = len(SB['asins']) - 1
        i = str(asin_idx[asin]); cur = SB['age'].get(i)
        if not cur or cur[1] == 0:
            SB['age'][i] = [round(charge, 2), int(units), (cur or [0, 0, 0, 0, 0, 0])[2], (cur or [0, 0, 0, 0, 0, 0])[3], (cur or [0, 0, 0, 0, 0, 0])[4], round(charge / units, 4) if units else 0]; nais += 1
    print('  AIS rows added:', nais)

# ---------- tracker ----------
tr = newest('Deal Tracker Updated*.xlsx')
dash = []
if tr:
    print('Tracker:', os.path.basename(tr)); wb = openpyxl.load_workbook(tr, read_only=True, data_only=True)
    for r in rows(wb['Cost Master'], 5):
        if len(r) >= 6 and r[1] and r[2] and num(r[3]) is not None: asin, sku = asin_sku(r[1], r[2]); upsert_cost(s(r[0]), asin, sku, num(r[3]), num(r[4]), num(r[5]))
    for r in rows(wb['Deal Dashboard'], 5):
        if len(r) < 27 or not hasattr(r[0], 'date'): continue
        date = r[0].date().isoformat()
        if date < A.since: continue
        asin, sku = s(r[16]), s(r[17])
        if not asin and not sku: continue
        dash.append({'date': date, 'tag': s(r[1]), 'deal': s(r[2]), 'asin': asin, 'sku': sku, 'ref': num(r[18]), 'max_deal': num(r[19]), 'final': num(r[20]), 'your_price': num(r[21]), 'min_price': num(r[22]), 'disc': num(r[23]), 'stock': num(r[24]), 'min_commit': num(r[25]), 'status': s(r[26]), 'action': s(r[27]) if len(r) > 27 else None, 'manual': num(r[15])})
    print('  dashboard rows since', A.since, ':', len(dash))
emit('dashboard', {'rows': dash, 'since': A.since, 'source': os.path.basename(tr) if tr else None})

# drop junk rows keyed by an ASIN instead of a SKU
before = len(SK); SK[:] = [x for x in SK if not (x.get('sku') and ASIN_RE.match(x['sku']) and not any(y for y in SK if y is not x and y.get('asin') == x['sku']))]
SK[:] = [x for x in SK if not (x.get('sku') and ASIN_RE.match(x['sku']))]; print('junk cost rows dropped', before - len(SK))
# ---------- re-tag Sellerboard ASINs ----------
retag = 0
for a in SB['asins']:
    if a[2] in (None, '?', ''):
        row = by_sku.get(a[1]) or by_asin.get(a[0])
        if row and row.get('tag'): a[2] = row['tag']; retag += 1
print('cost rows added', added, 'updated', updated, '| Sellerboard ASINs re-tagged', retag)
emit('skus', SK); emit('sellerboard', SB); emit('price_history', PH)
